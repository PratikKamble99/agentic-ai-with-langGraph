import requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TOKEN = "*****************************"
FILE_KEY = "******************"
HEADERS = {"X-Figma-Token": TOKEN}

# Get pages
file_data = requests.get(f"https://api.figma.com/v1/files/{FILE_KEY}?depth=1", headers=HEADERS).json()
pages = file_data["document"]["children"]

def extract_headings(node, page_name, results, depth=0):
    if node.get("type") == "TEXT":
        style = node.get("style", {})
        name = node.get("name", "")
        chars = node.get("characters", "")
        font_size = style.get("fontSize")
        font_weight = style.get("fontWeight")
        font_family = style.get("fontFamily")
        is_heading = (
            any(k in name.lower() for k in ["heading","title","h1","h2","h3","h4","section","hero","banner"])
            or (font_size and font_size >= 20)
        )
        if is_heading:
            results.append({
                "Page": page_name,
                "Layer Name": name,
                "Text Content": chars[:80] if chars else "",
                "Font Family": font_family,
                "Font Size": font_size,
                "Font Weight": font_weight,
            })
    for child in node.get("children", []):
        extract_headings(child, page_name, results, depth+1)

all_headings = []
for page in pages:
    page_data = requests.get(
        f"https://api.figma.com/v1/files/{FILE_KEY}/nodes?ids={page['id']}", headers=HEADERS
    ).json()
    node = page_data["nodes"][page["id"]]["document"]
    extract_headings(node, page["name"], all_headings)

# Create Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Section Headings"

headers = ["Page", "Layer Name", "Text Content", "Font Family", "Font Size (px)", "Font Weight"]
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="2D3748")
for i, h in enumerate(headers, 1):
    cell = ws.cell(1, i, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_i, r in enumerate(all_headings, 2):
    ws.cell(row_i, 1, r["Page"])
    ws.cell(row_i, 2, r["Layer Name"])
    ws.cell(row_i, 3, r["Text Content"])
    ws.cell(row_i, 4, r["Font Family"])
    ws.cell(row_i, 5, r["Font Size"])
    ws.cell(row_i, 6, r["Font Weight"])
    if row_i % 2 == 0:
        for c in range(1, 7):
            ws.cell(row_i, c).fill = PatternFill("solid", start_color="F7FAFC")

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 15

wb.save("figma_headings.xlsx")
print(f"Done! {len(all_headings)} headings extracted.")
# ```

# Install deps first: `pip install requests openpyxl`

# ---

### Option B — Share the API response here
# Open this URL in your browser (already logged in with your token via the URL), copy-paste the JSON response here, and I'll process it into an Excel file:
# ```
# https://api.figma.com/v1/files/4VxV2NPYFpoWWoDS84jV77?depth=2